import datetime
import openpyxl
import pandas as pd
import json
import archivator


def update_data(df: pd.DataFrame) -> None:
    archivator.screen()

    book = openpyxl.load_workbook("../../private/m2023.xlsx")

    uniq_months = list(map(int, df['month'].unique()))
    print(uniq_months)
    if len([i for i in uniq_months if i > len(book.sheetnames)]) != 0:
        create_new_sheets(df=df, book=book)

    for i in uniq_months:
        book.active = i - 1
        sheet = book.active
        df.query(f"month == '{i}'")
        for date, day, ethw, zil, kas, cons in zip(df['Date'], df['day'], df['ETHW reward'], df['ZIL reward'],
                                                   df['KAS reward'],
                                                   df['consumption']):

            if pd.Timestamp(sheet[f"A{day + 1}"].value) == pd.Timestamp(date):
                sheet[f"C{day + 1}"].value = ethw.replace('.', ',')
                sheet[f"D{day + 1}"].value = zil.replace('.', ',')
                sheet[f"E{day + 1}"].value = kas
                sheet[f"J{day + 1}"].value = cons * 1000

    book.save("../../private/m2023.xlsx")

    with open('../../private/last_update.json') as f:
        json_obj = json.load(f)
    json_obj['day'] = datetime.datetime.now().date().day - 1
    json_obj['month'] = datetime.datetime.now().date().month
    with open('../../private/last_update.json', 'w') as f:
        json.dump(json_obj, f)

    book.close()


def create_new_sheets(df: pd.DataFrame, book: openpyxl.Workbook):
    months = {1: 'January', 2: 'February', 3: 'March', 4: 'April ', 5: 'May ', 6: 'June ', 7: 'July ', 8: 'August ',
              9: 'September ', 10: 'October ', 11: 'November ', 12: 'December '}
    required_months = [i for i in df['month'].unique() if i > len(book.sheetnames)]
    for m in required_months:
        book.active = -1
        book.copy_worksheet(book.active)
        book.active = -1
        book.active.title = months[int(m)]
        sheet = book.active
        for r in range(2, 33):
            date = datetime.datetime.strptime(f"{r - 1}.{m}.2023", "%d.%m.%Y")

            if m in (1, 3, 5, 7, 8, 10, 12):
                sheet[f"A{r}"].value = date.date()
                sheet[f"A{r}"].number_format = "dd/mm/yyyy"
            else:
                if r < 32:
                    sheet[f"A{r}"].value = date.date()
                    sheet[f"A{r}"].number_format = "dd/mm/yyyy"

            for c in ['C', 'D', 'E', 'F', 'G', 'H', 'J']:
                sheet[f"{c}{r}"].value = None
    book.save('../../private/m2023.xlsx')
